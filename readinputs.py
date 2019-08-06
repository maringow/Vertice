# Read user input Excel file
import openpyxl as xl
import pandas as pd
import numpy as np

def read_model_inputs(parameters):
    wb = xl.load_workbook(filename=parameters['excel_filepath'], read_only=True, data_only=True) #data_only so it doesn't return the formulas
    sheet = wb['Input']

    # assign single-value variables from Excel cells into parameters dictionary
    parameters.update({'brand_status': sheet['B6'].value,
                       'channel': sheet['B7'].value,
                       'channel_detail': sheet['B8'].value,
                       'internal_external': sheet['B9'].value,
                       'vertice_filing_month': sheet['B10'].value,
                       'vertice_filing_year': sheet['B11'].value,
                       'vertice_launch_month': sheet['B12'].value,
                       'vertice_launch_year': sheet['B13'].value,
                       'indication': sheet['B14'].value,
                       'presentation': sheet['B15'].value,
                       'loe_year': sheet['B16'].value,
                       'competition_detail': sheet['B17'].value,
                       'pos': sheet['B18'].value,
                       'comments': sheet['B19'].value,
                       'scan_and_save': sheet['B20'].value,
                       'volume_growth_rate': sheet['B23'].value,
                       'wac_increase': sheet['B24'].value,
                       'gtn_%': sheet['B27'].value,
                       'DIO': sheet['B43'].value,
                       'DSO': sheet['B44'].value,
                       'DPO': sheet['B45'].value,
                       'discount_rate': sheet['B49'].value,
                       'tax_rate': sheet['B50'].value,
                       'exit_multiple': sheet['B51'].value,
                       'cogs': {'excipients': sheet['B31'].value,
                                'direct_labor': sheet['B32'].value,
                                'variable_overhead': sheet['B33'].value,
                                'fixed_overhead': sheet['B34'].value,
                                'depreciation': sheet['B35'].value,
                                'cmo_markup': sheet['B36'].value,
                                'cost_increase': sheet['B37'].value,
                                'distribution': sheet['B38'].value,
                                'writeoffs': sheet['B39'].value},
                       'present_year': sheet['B55'].value,
                       'last_forecasted_year': sheet['M55'].value
                       })

    # Set up df_gfm data frame
    df_gfm = pd.DataFrame()
    df_gfm['Year'] = list(range(2016, parameters['last_forecasted_year'] + 1, 1))
    df_gfm = df_gfm.set_index('Year')

    # Add excel yearly data
    def pull_yearly_data(row_number):  # row you want data from
        x = [0] * (parameters['present_year'] - 2016)  # zeros for years not in 'model input' excel sheet
        for i in range(2, 14):
            x.append(sheet.cell(row=row_number, column=i).value)
        return (x)

    df_gfm['Gx Penetration'] = pull_yearly_data(56)
    df_gfm['Number of Gx Players'] = pull_yearly_data(57)
    df_gfm['Vertice Gx Market Share'] = pull_yearly_data(58)
    df_gfm['Price Discount of Current Gx Net Price'] = pull_yearly_data(59)
    df_gfm['Profit Share %'] = pull_yearly_data(60)
    df_gfm['Milestone Payments'] = pull_yearly_data(61)
    df_gfm['SG&A'] = pull_yearly_data(62)
    df_gfm['R&D'] = pull_yearly_data(66)
    df_gfm['Total Capitalized'] = pull_yearly_data(75)
    df_gfm['Tax depreciation'] = pull_yearly_data(76)
    df_gfm['Additional Impacts on P&L'] = pull_yearly_data(84)
    df_gfm['Net proceeds from Disposals'] = pull_yearly_data(85)
    df_gfm['Write-off of Residual Tax Value'] = pull_yearly_data(86)
    df_gfm['Other Income, Expenses, Except Items'] = pull_yearly_data(87)
    df_gfm['Additional Non-cash Effects'] = pull_yearly_data(88)
    df_gfm['Other Net Current Assets'] = pull_yearly_data(89)
    df_gfm['Capital Avoidance'] = pull_yearly_data(90)
    df_gfm = df_gfm.fillna(0)  # if there is no data entered in the excel file, it gives NaNs, this converts them to 0s

    # Adding analog data
    sheet = wb['Analog']

    def pull_analog_data(row_number):  # row you want data from
        x = []
        for i in range(2, 12):
            x.append(sheet.cell(row=row_number, column=i).value)
        return (x)

    df_analog = pd.DataFrame(index=range(0, 10))
    df_analog['Retail Net Price Pct BWAC'] = pull_analog_data(2)
    df_analog['Clinic Net Price Pct BWAC'] = pull_analog_data(4)
    df_analog['Hospital Net Price Pct BWAC'] = pull_analog_data(6)

    df_analog['Clinic Market Share'] = pull_analog_data(5)
    df_analog['Retail Market Share'] = pull_analog_data(3)
    df_analog['Hospital Market Share'] = pull_analog_data(7)

    df_analog.index.name = "Number of Gx Players"
    df_analog = df_analog.fillna(0)

    return (parameters, df_gfm, df_analog)